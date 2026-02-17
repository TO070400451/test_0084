"""Excel reading utilities with header auto-detection."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.normalizer import normalize_cell_text


def _detect_header_row(
    ws: Worksheet,
    required_headers: list[str],
    max_scan: int = 200,
) -> tuple[int, dict[str, int]]:
    """Scan rows 1..max_scan for a row containing all required_headers.

    Returns (header_row_number, {header_name: column_index}).
    Column index is 1-based (openpyxl convention).
    """
    for row_idx in range(1, max_scan + 1):
        row_values: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1 if ws.max_column else 100):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                row_values[str(cell_val).strip()] = col_idx
        if all(h in row_values for h in required_headers):
            return row_idx, row_values
    raise ValueError(
        f"Header row not found within first {max_scan} rows. "
        f"Required headers: {required_headers}"
    )


def read_test_items(
    xlsx_path: str | Path,
    sheet_name: str = "Test Items",
) -> tuple[list[dict[str, str]], dict[str, int]]:
    """Read Test Items sheet and return list of row dicts + header map.

    Header detection: looks for row containing Test ID, Test Procedure, Check item.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[sheet_name]

    required = ["Test ID", "Test Procedure", "Check item"]
    header_row, header_map = _detect_header_row(ws, required, max_scan=50)

    # Columns to extract
    columns_of_interest = [
        "Test ID", "Section", "Sub-section", "Test Title",
        "Pre-Condition", "Test Procedure", "Check item",
        "Remark", "チーム分担",
    ]
    col_indices: dict[str, int] = {}
    for col_name in columns_of_interest:
        if col_name in header_map:
            col_indices[col_name] = header_map[col_name]

    rows: list[dict[str, str]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1 if ws.max_row else 10000):
        test_id_col = col_indices.get("Test ID")
        if test_id_col is None:
            break
        test_id_val = ws.cell(row=row_idx, column=test_id_col).value
        if test_id_val is None or str(test_id_val).strip() == "":
            continue
        row_data: dict[str, str] = {}
        for col_name, col_idx in col_indices.items():
            raw = ws.cell(row=row_idx, column=col_idx).value
            row_data[col_name] = normalize_cell_text(raw)
        rows.append(row_data)

    wb.close()
    return rows, header_map


def read_shikenkomoku_test_ids(
    xlsx_path: str | Path,
    sheet_name: str = "試験項目",
) -> list[str]:
    """Read the Test IDs from the Japanese Excel's 試験項目 sheet."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[sheet_name]

    required = ["No.", "Test ID", "Test Title"]
    header_row, header_map = _detect_header_row(ws, required, max_scan=200)

    test_id_col = header_map["Test ID"]
    ids: list[str] = []
    for row_idx in range(header_row + 1, ws.max_row + 1 if ws.max_row else 10000):
        val = ws.cell(row=row_idx, column=test_id_col).value
        if val is not None and str(val).strip():
            ids.append(str(val).strip())

    wb.close()
    return ids
