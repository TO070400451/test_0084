"""CLI: Apply patch.yml to Japanese Excel (試験項目 sheet).

Usage:
    python -m app.cli_patcher \
        --base "input/master.xlsx" \
        --patch "out/patch.yml" \
        --sheet "試験項目" \
        --output "out/master_updated.xlsx" \
        --report "out/diff.md"
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

from app.diff_report import generate_diff_report
from app.excel_write import apply_patch
from app.patch_io import read_patch
from app.renumber import renumber_sheet


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Apply patch.yml to Japanese Excel 試験項目 sheet"
    )
    parser.add_argument(
        "--base", required=True, help="Path to base Japanese Excel"
    )
    parser.add_argument(
        "--patch", required=True, help="Path to patch.yml"
    )
    parser.add_argument(
        "--sheet", default="試験項目", help="Target sheet name"
    )
    parser.add_argument(
        "--output", required=True, help="Output Excel path"
    )
    parser.add_argument(
        "--report", default="out/diff.md", help="Diff report output path"
    )
    parser.add_argument(
        "--end-empty-rows", type=int, default=3,
        help="Consecutive empty rows to detect data end"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Only generate diff report without writing Excel"
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)

    # 1. Read patch
    print(f"Reading patch: {args.patch}")
    patch = read_patch(args.patch)
    patch.sheet = args.sheet
    print(f"  Operations: {len(patch.operations)}")

    if args.dry_run:
        print("Dry run mode: skipping Excel write.")
        # TODO: Implement dry-run diff
        return

    # 2. Apply patch
    print(f"Applying patch to: {args.base}")
    diff_entries = apply_patch(
        args.base, patch, args.output,
        end_empty_rows=args.end_empty_rows,
    )

    # 3. Renumber No. column
    print(f"Renumbering No. column...")
    wb = openpyxl.load_workbook(str(args.output))
    ws = wb[args.sheet]

    # Detect header for renumbering
    header_row = None
    no_col = None
    test_id_col = None
    for row_idx in range(1, 201):
        found_no = False
        found_tid = False
        for col_idx in range(1, ws.max_column + 1 if ws.max_column else 100):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val).strip()
                if s == "No.":
                    no_col = col_idx
                    found_no = True
                elif s == "Test ID":
                    test_id_col = col_idx
                    found_tid = True
        if found_no and found_tid:
            header_row = row_idx
            break

    if header_row and no_col and test_id_col:
        count = renumber_sheet(
            ws, header_row, no_col, test_id_col,
            end_empty_rows=args.end_empty_rows,
        )
        print(f"  Renumbered {count} rows.")
        wb.save(str(args.output))
    else:
        print("  Warning: Could not detect No./Test ID headers for renumbering.")

    wb.close()

    # 4. Generate diff report
    generate_diff_report(diff_entries, args.report)
    print(f"Report written: {args.report}")
    print(f"Output written: {args.output}")


if __name__ == "__main__":
    main()
