"""Auto-renumbering for No. column in 試験項目 sheet."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


def renumber_sheet(
    ws: Worksheet,
    header_row: int,
    no_col: int,
    test_id_col: int,
    *,
    end_empty_rows: int = 3,
) -> int:
    """Renumber the No. column (1, 2, 3, ...) for rows with non-empty Test ID.

    Args:
        ws: The worksheet to renumber.
        header_row: The header row number.
        no_col: Column index of "No." (1-based).
        test_id_col: Column index of "Test ID" (1-based).
        end_empty_rows: Number of consecutive empty Test ID rows to detect end of data.

    Returns:
        Total count of numbered rows.
    """
    counter = 0
    empty_streak = 0

    row_idx = header_row + 1
    max_row = ws.max_row or (header_row + 10000)

    while row_idx <= max_row:
        test_id_val = ws.cell(row=row_idx, column=test_id_col).value
        if test_id_val is not None and str(test_id_val).strip():
            empty_streak = 0
            counter += 1
            ws.cell(row=row_idx, column=no_col).value = counter
        else:
            empty_streak += 1
            if empty_streak >= end_empty_rows:
                break
        row_idx += 1

    return counter
