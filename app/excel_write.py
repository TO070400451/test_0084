"""Excel writing utilities for the patcher (update/insert with format preservation)."""

from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.patch_model import InsertOperation, PatchFile, UpdateOperation

# Headers that should never be overwritten
_PROTECTED_HEADER_PATTERNS = [
    "自動入力",
    "TestNo",
    "TestIDの試験数",
]

# Model name columns (should not be overwritten)
_MODEL_PATTERN = re.compile(r"^EB\d{4}$")


def _is_protected_header(header: str) -> bool:
    """Check if a header name is protected (should not be overwritten)."""
    for pattern in _PROTECTED_HEADER_PATTERNS:
        if pattern in header:
            return True
    if _MODEL_PATTERN.match(header):
        return True
    return False


def _detect_header_row(
    ws: Worksheet,
    required_headers: list[str],
    max_scan: int = 200,
) -> tuple[int, dict[str, int]]:
    """Same as excel_read but for write context (not read_only)."""
    for row_idx in range(1, max_scan + 1):
        row_values: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1 if ws.max_column else 100):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                row_values[str(cell_val).strip()] = col_idx
        if all(h in row_values for h in required_headers):
            return row_idx, row_values
    raise ValueError(
        f"Header row not found within first {max_scan} rows. "
        f"Required: {required_headers}"
    )


def _find_row_by_test_id(
    ws: Worksheet,
    test_id: str,
    header_row: int,
    test_id_col: int,
    *,
    end_empty_rows: int = 3,
) -> int | None:
    """Find the row number for a given Test ID."""
    empty_streak = 0
    max_row = ws.max_row or (header_row + 10000)
    for row_idx in range(header_row + 1, max_row + 1):
        val = ws.cell(row=row_idx, column=test_id_col).value
        if val is not None and str(val).strip() == test_id:
            return row_idx
        if val is None or str(val).strip() == "":
            empty_streak += 1
            if empty_streak >= end_empty_rows:
                break
        else:
            empty_streak = 0
    return None


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    """Copy formatting from source cell to destination cell."""
    dst.font = copy.copy(src.font)
    dst.border = copy.copy(src.border)
    dst.fill = copy.copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy.copy(src.protection)
    dst.alignment = copy.copy(src.alignment)


def _copy_cell_formula_or_clear(src: Cell, dst: Cell) -> None:
    """Copy formula from source (if it has one), otherwise leave value to be set later."""
    if isinstance(src.value, str) and src.value.startswith("="):
        dst.value = src.value
    # Otherwise leave dst.value as None (to be set by patch data)


def _insert_row_after(
    ws: Worksheet,
    after_row: int,
    header_map: dict[str, int],
    row_data: dict[str, str],
) -> int:
    """Insert a row after the given row, copying style from after_row as template.

    Returns the new row number.
    """
    new_row = after_row + 1
    ws.insert_rows(new_row)

    # Copy style and formulas from the template row (after_row)
    max_col = ws.max_column or 50
    for col_idx in range(1, max_col + 1):
        src_cell = ws.cell(row=after_row, column=col_idx)
        dst_cell = ws.cell(row=new_row, column=col_idx)
        _copy_cell_style(src_cell, dst_cell)
        _copy_cell_formula_or_clear(src_cell, dst_cell)

    # Copy row height
    if after_row in ws.row_dimensions:
        src_rd = ws.row_dimensions[after_row]
        dst_rd = ws.row_dimensions[new_row]
        dst_rd.height = src_rd.height

    # Set values from patch data
    for col_name, value in row_data.items():
        if col_name in header_map:
            col_idx = header_map[col_name]
            # Don't overwrite protected columns unless explicitly in row_data
            ws.cell(row=new_row, column=col_idx).value = value

    return new_row


def apply_patch(
    xlsx_path: str | Path,
    patch: PatchFile,
    output_path: str | Path,
    *,
    end_empty_rows: int = 3,
) -> list[dict[str, Any]]:
    """Apply a patch to an Excel file and save to output_path.

    Returns a list of diff entries for reporting.
    """
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb[patch.sheet]

    required = ["No.", "Test ID", "Test Title"]
    header_row, header_map = _detect_header_row(ws, required)
    test_id_col = header_map["Test ID"]

    # Identify protected columns
    protected_cols: set[int] = set()
    for name, col_idx in header_map.items():
        if _is_protected_header(name):
            protected_cols.add(col_idx)

    diff_entries: list[dict[str, Any]] = []

    for op in patch.operations:
        if isinstance(op, UpdateOperation):
            row_num = _find_row_by_test_id(
                ws, op.test_id, header_row, test_id_col,
                end_empty_rows=end_empty_rows,
            )
            if row_num is None:
                diff_entries.append({
                    "type": "warning",
                    "test_id": op.test_id,
                    "message": "Test ID not found for update; skipped.",
                })
                continue

            entry: dict[str, Any] = {
                "type": "update",
                "test_id": op.test_id,
                "changes": {},
            }
            for col_name, new_val in op.set_values.items():
                if col_name not in header_map:
                    continue
                col_idx = header_map[col_name]
                if col_idx in protected_cols:
                    continue
                old_val = ws.cell(row=row_num, column=col_idx).value
                ws.cell(row=row_num, column=col_idx).value = new_val
                entry["changes"][col_name] = {
                    "old": str(old_val) if old_val else "",
                    "new": str(new_val),
                }
            diff_entries.append(entry)

        elif isinstance(op, InsertOperation):
            after_row = _find_row_by_test_id(
                ws, op.after_test_id, header_row, test_id_col,
                end_empty_rows=end_empty_rows,
            )
            if after_row is None:
                diff_entries.append({
                    "type": "warning",
                    "test_id": op.row.get("Test ID", "?"),
                    "message": f"after_key '{op.after_test_id}' not found; skipped.",
                })
                continue

            new_row_num = _insert_row_after(ws, after_row, header_map, op.row)
            diff_entries.append({
                "type": "insert",
                "test_id": op.row.get("Test ID", "?"),
                "after_key": op.after_test_id,
                "row_num": new_row_num,
            })

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    return diff_entries
